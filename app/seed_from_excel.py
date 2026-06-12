"""
One-time seed: imports all data from the V11 Excel workbook into Google Sheets.
Run this once to populate the sheets.

V11 sheet → Google Sheets table mapping:
  A3. RM Master        → raw_materials
  A4. Vendor Rate Master → raw_materials (rates / MOQ / lead time)
  A1. SKU Master       → products
  A2. Formulation Master → formulations
  B1. Launch Plan      → skus
  A5. PM Component Master → packaging_configs
  A3. RM Master col G  → inventory
  (hardcoded)          → launch_phases
"""

import os
import sys
import time
import openpyxl
import gspread

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Inovae_NutraCare_V11.xlsx')
CREDS_FILE = os.path.join(os.path.dirname(__file__), '..', 'credentials.json')
SHEET_ID = "1HAYm6Jm1_Inakqv04krAruHUQbryZoTff7OQckMAZsA"

SCHEMAS = {
    'raw_materials': ['id', 'name', 'moq_kg', 'rate_per_kg', 'category', 'source',
                      'negotiator', 'vendor1', 'vendor2', 'vendor3', 'remarks'],
    'products': ['id', 'base_code', 'category', 'delivery_form', 'daily_consumption',
                 'units_per_pack', 'processing_cost_per_unit', 'packaging_cost_per_pack',
                 'clinical_trial_required', 'clinical_duration_months', 'clinical_participants',
                 'clinical_packs_per_month', 'margin_percent', 'buffer_percent'],
    'formulations': ['id', 'product_id', 'raw_material_id', 'grams_per_unit'],
    'skus': ['id', 'product_id', 'flavour_code', 'flavour_name', 'sku_code', 'pack_count', 'phase'],
    'packaging_configs': ['id', 'delivery_form', 'component_name', 'cost_per_pack'],
    'inventory': ['id', 'raw_material_id', 'qty_grams', 'location'],
    'launch_phases': ['id', 'phase_number', 'start_date', 'gap_days'],
}


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def safe_float(v, default=0):
    if v is None or v == '':
        return default
    if isinstance(v, str):
        try:
            return float(v.replace(',', ''))
        except ValueError:
            return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def batch_write(ws, rows, schema):
    """Write rows in batches to respect API rate limits."""
    BATCH_SIZE = 50
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        values = []
        for row in batch:
            values.append([safe_str(row.get(col, '')) for col in schema])
        start_row = i + 2  # +1 header, +1 for 1-indexed
        end_col = chr(ord('A') + len(schema) - 1)
        ws.update(
            values=values,
            range_name=f'A{start_row}:{end_col}{start_row + len(values) - 1}',
            value_input_option='RAW',
        )
        if i + BATCH_SIZE < len(rows):
            time.sleep(1)


def seed():
    print("Connecting to Google Sheets...")
    gc = gspread.service_account(filename=CREDS_FILE)
    ss = gc.open_by_key(SHEET_ID)

    # ─── Create / clear all worksheets ────────────────────────────────────────
    existing = {ws.title: ws for ws in ss.worksheets()}
    for table_name, schema in SCHEMAS.items():
        if table_name in existing:
            ws = existing[table_name]
            ws.clear()
            ws.update(values=[schema], range_name='A1')
        else:
            ws = ss.add_worksheet(title=table_name, rows=1000, cols=len(schema))
            ws.update(values=[schema], range_name='A1')
        ws.format('A1:Z1', {'textFormat': {'bold': True}})
        time.sleep(0.5)

    # Remove default Sheet1 if present
    if 'Sheet1' in existing:
        try:
            ss.del_worksheet(existing['Sheet1'])
        except Exception:
            pass

    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # =========================================================================
    # 1. RAW MATERIALS
    #    A3. RM Master: A=RM ID, B=RM Name, C=Category, D=UOM, E=Regulatory/Source,
    #                   F=Active Status, G=Avail @ Karsy (g)
    #    A4. Vendor Rate Master: A=RM ID, F=Rate/kg, G=MOQ, H=Lead Time
    # =========================================================================
    print("Importing raw materials...")
    ws_rm = wb['A3. RM Master']
    ws_vr = wb['A4. Vendor Rate Master']

    # Build vendor-rate lookup keyed by RM ID (integer)
    vendor_rates = {}  # rm_id -> {rate, moq, lead_time, vendor_name, negotiator}
    for row in range(4, ws_vr.max_row + 1):
        rm_id_raw = ws_vr.cell(row=row, column=1).value
        if rm_id_raw is None:
            continue
        try:
            rm_id_key = int(rm_id_raw)
        except (ValueError, TypeError):
            continue
        active = safe_str(ws_vr.cell(row=row, column=10).value).lower()
        if active != 'yes':
            continue
        vendor_rates[rm_id_key] = {
            'vendor_name': safe_str(ws_vr.cell(row=row, column=4).value),
            'negotiator': safe_str(ws_vr.cell(row=row, column=5).value),
            'rate': safe_float(ws_vr.cell(row=row, column=6).value),
            'moq': safe_float(ws_vr.cell(row=row, column=7).value),
            'lead_time': safe_float(ws_vr.cell(row=row, column=8).value),
        }

    rm_rows = []
    rm_id_to_seq = {}   # excel RM ID (int) -> sequential row id
    rm_id_to_name = {}  # excel RM ID (int) -> name string

    seq_id = 0
    for row in range(4, ws_rm.max_row + 1):
        rm_id_raw = ws_rm.cell(row=row, column=1).value
        name_raw = ws_rm.cell(row=row, column=2).value
        if rm_id_raw is None or name_raw is None:
            continue
        name = safe_str(name_raw)
        if not name:
            continue
        try:
            rm_id_excel = int(rm_id_raw)
        except (ValueError, TypeError):
            continue

        seq_id += 1
        rm_id_to_seq[rm_id_excel] = seq_id
        rm_id_to_name[rm_id_excel] = name

        category = safe_str(ws_rm.cell(row=row, column=3).value)
        source = safe_str(ws_rm.cell(row=row, column=5).value)

        vr = vendor_rates.get(rm_id_excel, {})
        rate = vr.get('rate', 0)
        moq = vr.get('moq', 0)
        lead_time = vr.get('lead_time', 0)
        vendor_name = vr.get('vendor_name', '')
        negotiator = vr.get('negotiator', '')

        # Store lead time in remarks (no dedicated column yet)
        remarks = f"Lead Time: {int(lead_time)} days" if lead_time else ''

        rm_rows.append({
            'id': seq_id,
            'name': name,
            'moq_kg': moq,
            'rate_per_kg': rate,
            'category': category,
            'source': source,
            'negotiator': negotiator,
            'vendor1': vendor_name,
            'vendor2': '',
            'vendor3': '',
            'remarks': remarks,
        })

    ws_gsheet = ss.worksheet('raw_materials')
    batch_write(ws_gsheet, rm_rows, SCHEMAS['raw_materials'])
    print(f"  Imported {len(rm_rows)} raw materials")
    time.sleep(1)

    # =========================================================================
    # 2. PACKAGING COST per SKU from A6. Packaging BOM
    #    Columns: A=SKU ID, D=Qty per Pack, F=Rate (Rs), G=Cost per Pack (Rs)
    #    Sum cost-per-pack values for each SKU ID
    # =========================================================================
    print("Computing packaging costs from A6. Packaging BOM...")
    ws_bom = wb['A6. Packaging BOM']
    pkg_cost_by_sku = {}  # sku_id -> total packaging cost per pack

    for row in range(4, ws_bom.max_row + 1):
        sku_id_raw = ws_bom.cell(row=row, column=1).value
        if sku_id_raw is None:
            continue
        sku_id_str = safe_str(sku_id_raw)
        if not sku_id_str:
            continue
        cost_per_pack = safe_float(ws_bom.cell(row=row, column=7).value)
        pkg_cost_by_sku[sku_id_str] = pkg_cost_by_sku.get(sku_id_str, 0) + cost_per_pack

    print(f"  Packaging cost map: {len(pkg_cost_by_sku)} SKUs")

    # =========================================================================
    # 3. PRODUCTS
    #    A1. SKU Master: A=SKU ID, B=Display Name, C=Category, D=Format,
    #    E=Status, F=Flavour, G=Pack Count, H=Units/Pack, I=Daily Dose,
    #    J=Grammage (formula→data_only), K=Buffer%, L=GM% Target,
    #    M=Proc Cost/Unit
    #    base_code = col A, category = col C, delivery_form = col D
    # =========================================================================
    print("Importing products...")
    ws_sku = wb['A1. SKU Master']
    prod_rows = []
    product_seq_id = 0
    sku_id_to_prod_id = {}  # SKU ID string -> sequential product id

    for row in range(4, ws_sku.max_row + 1):
        sku_id_raw = ws_sku.cell(row=row, column=1).value
        if sku_id_raw is None:
            continue
        sku_id_str = safe_str(sku_id_raw)
        if not sku_id_str:
            continue

        category = safe_str(ws_sku.cell(row=row, column=3).value)
        delivery_form = safe_str(ws_sku.cell(row=row, column=4).value)
        units_per_pack = safe_float(ws_sku.cell(row=row, column=8).value, 30)
        daily_dose = safe_float(ws_sku.cell(row=row, column=9).value, 1)
        buffer_pct = safe_float(ws_sku.cell(row=row, column=11).value, 0.05)
        margin_pct = safe_float(ws_sku.cell(row=row, column=12).value, 0.70)
        proc_cost = safe_float(ws_sku.cell(row=row, column=13).value, 0)
        pkg_cost = pkg_cost_by_sku.get(sku_id_str, 0)

        product_seq_id += 1
        sku_id_to_prod_id[sku_id_str] = product_seq_id

        prod_rows.append({
            'id': product_seq_id,
            'base_code': sku_id_str,
            'category': category,
            'delivery_form': delivery_form,
            'daily_consumption': int(daily_dose),
            'units_per_pack': int(units_per_pack),
            'processing_cost_per_unit': proc_cost,
            'packaging_cost_per_pack': round(pkg_cost, 4),
            'clinical_trial_required': 0,
            'clinical_duration_months': 0,
            'clinical_participants': 0,
            'clinical_packs_per_month': 1,
            'margin_percent': margin_pct,
            'buffer_percent': buffer_pct,
        })

    ws_gsheet = ss.worksheet('products')
    batch_write(ws_gsheet, prod_rows, SCHEMAS['products'])
    print(f"  Imported {len(prod_rows)} products")
    time.sleep(1)

    # =========================================================================
    # 4. FORMULATIONS
    #    A2. Formulation Master: A=SKU ID, B=RM ID, C=Ingredient Name,
    #    D=Qty/Unit (g), E=UOM, F=Formulation Version, G=Active Version Flag
    #    Only import rows where G = "Yes"
    # =========================================================================
    print("Importing formulations...")
    ws_form = wb['A2. Formulation Master']
    form_rows = []
    form_seq_id = 0
    skipped_form = 0

    for row in range(4, ws_form.max_row + 1):
        sku_id_raw = ws_form.cell(row=row, column=1).value
        if sku_id_raw is None:
            continue
        sku_id_str = safe_str(sku_id_raw)
        if not sku_id_str:
            continue

        active_flag = safe_str(ws_form.cell(row=row, column=7).value)
        if active_flag.lower() != 'yes':
            skipped_form += 1
            continue

        rm_id_raw = ws_form.cell(row=row, column=2).value
        if rm_id_raw is None:
            skipped_form += 1
            continue
        try:
            rm_id_excel = int(rm_id_raw)
        except (ValueError, TypeError):
            skipped_form += 1
            continue

        grams = safe_float(ws_form.cell(row=row, column=4).value)
        if grams <= 0:
            skipped_form += 1
            continue

        prod_id = sku_id_to_prod_id.get(sku_id_str)
        if not prod_id:
            skipped_form += 1
            continue

        rm_seq = rm_id_to_seq.get(rm_id_excel)
        if not rm_seq:
            skipped_form += 1
            continue

        form_seq_id += 1
        form_rows.append({
            'id': form_seq_id,
            'product_id': prod_id,
            'raw_material_id': rm_seq,
            'grams_per_unit': grams,
        })

    ws_gsheet = ss.worksheet('formulations')
    batch_write(ws_gsheet, form_rows, SCHEMAS['formulations'])
    print(f"  Imported {len(form_rows)} formulation entries (skipped {skipped_form})")
    time.sleep(1)

    # =========================================================================
    # 5. SKUs
    #    B1. Launch Plan: A=SKU ID, B=Display Name, C=Category, D=Format,
    #    E=Status, F=Phase 1, G=Phase 2, H=Phase 3, I=Phase 4, J=Trial
    #    pack_count = sum of phases + trial
    #    phase = first non-zero phase column (1-4), else 1 if trial only
    # =========================================================================
    print("Importing SKUs...")
    ws_lp = wb['B1. Launch Plan']
    sku_rows = []
    sku_seq_id = 0

    for row in range(3, ws_lp.max_row + 1):
        sku_id_raw = ws_lp.cell(row=row, column=1).value
        if sku_id_raw is None:
            continue
        sku_id_str = safe_str(sku_id_raw)
        # Skip summary / totals rows
        if not sku_id_str or sku_id_str.upper() in ('SKU ID', 'TOTAL'):
            continue

        prod_id = sku_id_to_prod_id.get(sku_id_str)
        if not prod_id:
            print(f"  WARNING: No product match for SKU '{sku_id_str}' in launch plan — skipping")
            continue

        display_name = safe_str(ws_lp.cell(row=row, column=2).value)
        # Flavour from A1. SKU Master (already loaded), default empty
        flavour = ''
        for pr in prod_rows:
            if pr['base_code'] == sku_id_str:
                break
        # look up flavour from SKU master sheet
        flavour_val = ws_sku.cell(
            row=3 + list(sku_id_to_prod_id.keys()).index(sku_id_str) + 1,
            column=6
        ).value if sku_id_str in sku_id_to_prod_id else None
        # simpler: iterate sku sheet for flavour
        for sr in range(4, ws_sku.max_row + 1):
            if safe_str(ws_sku.cell(row=sr, column=1).value) == sku_id_str:
                flavour = safe_str(ws_sku.cell(row=sr, column=6).value)
                break

        phase_vals = [
            safe_float(ws_lp.cell(row=row, column=6).value),   # Phase 1
            safe_float(ws_lp.cell(row=row, column=7).value),   # Phase 2
            safe_float(ws_lp.cell(row=row, column=8).value),   # Phase 3
            safe_float(ws_lp.cell(row=row, column=9).value),   # Phase 4
            safe_float(ws_lp.cell(row=row, column=10).value),  # Trial
        ]
        pack_count = int(sum(phase_vals))

        # Determine first non-zero phase (1-4); if only trial → phase 1
        phase = 1
        for p_idx, p_val in enumerate(phase_vals[:4], start=1):
            if p_val > 0:
                phase = p_idx
                break

        sku_seq_id += 1
        sku_rows.append({
            'id': sku_seq_id,
            'product_id': prod_id,
            'flavour_code': flavour[:2] if flavour else '',
            'flavour_name': flavour,
            'sku_code': sku_id_str,
            'pack_count': pack_count,
            'phase': phase,
        })

    ws_gsheet = ss.worksheet('skus')
    batch_write(ws_gsheet, sku_rows, SCHEMAS['skus'])
    print(f"  Imported {len(sku_rows)} SKUs")
    time.sleep(1)

    # =========================================================================
    # 6. PACKAGING CONFIGS
    #    A5. PM Component Master: A=PM ID, B=Component Name,
    #    C=Applicable Formats, D=UOM, E=Rate, F=MOQ
    # =========================================================================
    print("Importing packaging configs...")
    ws_pm = wb['A5. PM Component Master']
    pkg_rows = []
    pkg_seq_id = 0

    for row in range(4, ws_pm.max_row + 1):
        pm_id_raw = ws_pm.cell(row=row, column=1).value
        if pm_id_raw is None:
            continue
        pm_id_str = safe_str(pm_id_raw)
        if not pm_id_str:
            continue

        comp_name = safe_str(ws_pm.cell(row=row, column=2).value)
        formats_raw = safe_str(ws_pm.cell(row=row, column=3).value)
        rate = safe_float(ws_pm.cell(row=row, column=5).value)

        # Expand formats: "POW,GEL" → one row per format, "ALL" → one row with ALL
        formats = [f.strip() for f in formats_raw.split(',') if f.strip()] if formats_raw else ['ALL']

        for fmt in formats:
            pkg_seq_id += 1
            pkg_rows.append({
                'id': pkg_seq_id,
                'delivery_form': fmt,
                'component_name': comp_name,
                'cost_per_pack': rate,
            })

    ws_gsheet = ss.worksheet('packaging_configs')
    batch_write(ws_gsheet, pkg_rows, SCHEMAS['packaging_configs'])
    print(f"  Imported {len(pkg_rows)} packaging config entries")
    time.sleep(1)

    # =========================================================================
    # 7. INVENTORY
    #    A3. RM Master col G = Avail @ Karsy (g)
    #    Map to raw_material_id via rm_id_to_seq
    # =========================================================================
    print("Importing inventory...")
    inv_rows = []
    inv_seq_id = 0

    for row in range(4, ws_rm.max_row + 1):
        rm_id_raw = ws_rm.cell(row=row, column=1).value
        if rm_id_raw is None:
            continue
        try:
            rm_id_excel = int(rm_id_raw)
        except (ValueError, TypeError):
            continue

        qty_raw = ws_rm.cell(row=row, column=7).value  # col G = Avail @ Karsy (g)
        qty_grams = safe_float(qty_raw)
        if qty_grams <= 0:
            continue

        rm_seq = rm_id_to_seq.get(rm_id_excel)
        if not rm_seq:
            continue

        inv_seq_id += 1
        inv_rows.append({
            'id': inv_seq_id,
            'raw_material_id': rm_seq,
            'qty_grams': qty_grams,
            'location': 'Karsy',
        })

    ws_gsheet = ss.worksheet('inventory')
    batch_write(ws_gsheet, inv_rows, SCHEMAS['inventory'])
    print(f"  Imported {len(inv_rows)} inventory entries")
    time.sleep(1)

    # =========================================================================
    # 8. LAUNCH PHASES (4 static entries)
    # =========================================================================
    print("Importing launch phases...")
    phase_rows = [
        {'id': 1, 'phase_number': 1, 'start_date': '2025-01-01', 'gap_days': 0},
        {'id': 2, 'phase_number': 2, 'start_date': '2025-04-01', 'gap_days': 90},
        {'id': 3, 'phase_number': 3, 'start_date': '2025-07-01', 'gap_days': 90},
        {'id': 4, 'phase_number': 4, 'start_date': '2025-10-01', 'gap_days': 90},
    ]
    ws_gsheet = ss.worksheet('launch_phases')
    batch_write(ws_gsheet, phase_rows, SCHEMAS['launch_phases'])
    print("  Imported 4 launch phases")

    wb.close()
    print("\nDone! Google Sheet seeded successfully from V11.")
    print(f"View at: https://docs.google.com/spreadsheets/d/{SHEET_ID}")


if __name__ == '__main__':
    seed()
